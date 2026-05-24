"""Sync TrialDesignBench dataset from Google Sheets to local + Hugging Face.

Steps:
1. Download the latest sheet as CSV from Google Sheets.
2. Diff against the existing tdr.parquet by the "#" column to find new rows.
3. Download protocol and SAP PDFs for new rows (skip if no link).
4. Overwrite tdr.parquet and upload the changed files to Hugging Face.

Usage:
    python sync_dataset.py              # full sync
    python sync_dataset.py --no-upload  # local only
    python sync_dataset.py --dry-run    # show what would happen
"""

from __future__ import annotations

import argparse
import csv
import subprocess
import sys
import time
import urllib.error
import urllib.request
from pathlib import Path
from urllib.parse import urlparse

SHEET_ID = "1Vb6U9Jzigtg5hLcn4R_5REW_G84cNVk6"
SHEET_GID = "0"
SHEET_CSV_URL = (
    f"https://docs.google.com/spreadsheets/d/{SHEET_ID}/export"
    f"?format=csv&gid={SHEET_GID}"
)
# Fallback for uploaded .xlsx files (htmlview URLs) — downloads raw bytes
# and converts to CSV via openpyxl.
SHEET_XLSX_URL = f"https://docs.google.com/uc?export=download&id={SHEET_ID}"
SHEET_TAB_INDEX = 0  # 0 = first sheet; change if data is on another tab

# Data layout — overridden by --data-dir. Defaults to the script directory so a
# self-contained `source/` checkout still works without flags.
ROOT = Path(__file__).parent
PARQUET_PATH = ROOT / "data" / "tdr.parquet"
DOCS_DIR = ROOT / "documents"
HF_REPO = "trialdesignbench/source"


def _set_data_dir(data_dir: Path) -> None:
    """Repoint the module-level paths at a different data directory."""
    global ROOT, PARQUET_PATH, DOCS_DIR
    ROOT = data_dir.resolve()
    PARQUET_PATH = ROOT / "data" / "tdr.parquet"
    DOCS_DIR = ROOT / "documents"

UA = (
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
    "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
)
TIMEOUT = 60
RETRIES = 2


def fetch_sheet_csv() -> str:
    """Try native-sheet CSV export first; fall back to xlsx download + convert."""
    try:
        req = urllib.request.Request(SHEET_CSV_URL, headers={"User-Agent": UA})
        with urllib.request.urlopen(req, timeout=TIMEOUT) as resp:
            final_url = resp.geturl()
            if "accounts.google.com" in final_url or "ServiceLogin" in final_url:
                raise RuntimeError(
                    "Google Sheet is not publicly shared — "
                    "enable 'Anyone with link: Viewer'."
                )
            return resp.read().decode("utf-8")
    except urllib.error.HTTPError as e:
        if e.code != 400:
            raise
        print("CSV export returned 400 — falling back to .xlsx download.")
        return _fetch_xlsx_as_csv()


def _fetch_xlsx_as_csv() -> str:
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        msg = (
            "openpyxl is required to read uploaded .xlsx Drive files. "
            "Install with: pip install openpyxl"
        )
        raise RuntimeError(msg) from e

    import io

    req = urllib.request.Request(SHEET_XLSX_URL, headers={"User-Agent": UA})
    with urllib.request.urlopen(req, timeout=TIMEOUT) as resp:
        final_url = resp.geturl()
        if "accounts.google.com" in final_url or "ServiceLogin" in final_url:
            raise RuntimeError(
                "File is not publicly shared — enable 'Anyone with link: Viewer'."
            )
        data = resp.read()

    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.worksheets[SHEET_TAB_INDEX]

    def _fmt(v: object) -> str:
        if v is None:
            return ""
        # openpyxl returns whole-number cells as floats (1.0, 2.0, ...);
        # collapse back to int so "#" / "Year" / "PMID" stay digit-like.
        if isinstance(v, float) and v.is_integer():
            return str(int(v))
        return str(v)

    buf = io.StringIO()
    writer = csv.writer(buf)
    for row in ws.iter_rows(values_only=True):
        writer.writerow([_fmt(v) for v in row])
    return buf.getvalue()


def read_existing_ids() -> set[str]:
    if not PARQUET_PATH.exists():
        return set()
    import pandas as pd

    df = pd.read_parquet(PARQUET_PATH, columns=["#"])
    out: set[str] = set()
    for v in df["#"].dropna():
        # Normalize whole-number floats ("1.0") back to "1" so the diff
        # matches sheet rows regardless of how the column was stored.
        if isinstance(v, float) and v.is_integer():
            out.add(str(int(v)))
        else:
            s = str(v).strip()
            if s:
                out.add(s)
    return out


def write_parquet(csv_text: str) -> None:
    import io

    import pandas as pd

    df = pd.read_csv(io.StringIO(csv_text))
    PARQUET_PATH.parent.mkdir(parents=True, exist_ok=True)
    df.to_parquet(PARQUET_PATH, index=False, compression="snappy")


def parse_rows(csv_text: str) -> list[dict[str, str]]:
    reader = csv.DictReader(csv_text.splitlines())
    return [row for row in reader if (row.get("#") or "").strip().isdigit()]


def download_pdf(url: str, dest: Path) -> tuple[bool, str]:
    if dest.exists() and dest.stat().st_size > 0:
        return True, "exists"
    req = urllib.request.Request(url, headers={"User-Agent": UA})
    last_err = ""
    for attempt in range(RETRIES + 1):
        try:
            with urllib.request.urlopen(req, timeout=TIMEOUT) as resp:
                data = resp.read()
            dest.write_bytes(data)
            return True, f"ok ({len(data)} bytes)"
        except urllib.error.HTTPError as e:
            last_err = f"HTTP {e.code}"
        except (urllib.error.URLError, TimeoutError) as e:
            last_err = f"network: {e}"
        except Exception as e:  # noqa: BLE001
            last_err = f"error: {e}"
        time.sleep(1 + attempt)
    return False, last_err


def paper_link_slug(link: str) -> str:
    """Last two path segments of a Paper Link joined with '_'.

    Example: https://doi.org/10.1056/nejmoa2511478 -> 10.1056_nejmoa2511478
    """
    path = urlparse((link or "").strip()).path.strip("/")
    parts = [p for p in path.split("/") if p]
    return "_".join(parts[-2:]) if len(parts) >= 2 else ""


def _row_missing_pdfs(row: dict[str, str]) -> bool:
    """A row needs work iff it has a usable slug + at least one link whose
    target PDF is not already on disk."""
    slug = paper_link_slug(row.get("Paper Link") or "")
    if not slug:
        return False
    protocol = (row.get("Study Protocol Link") or "").strip()
    sap = (row.get("Protocol+SAP / SAP Link") or "").strip()
    if not protocol and not sap:
        return False
    row_dir = DOCS_DIR / slug
    if protocol and not (row_dir / "protocol.pdf").exists():
        return True
    if sap and not (row_dir / "sap.pdf").exists():
        return True
    return False


def download_for_row(row: dict[str, str]) -> list[Path]:
    num = (row.get("#") or "").strip()
    slug = paper_link_slug(row.get("Paper Link") or "")
    protocol = (row.get("Study Protocol Link") or "").strip()
    sap = (row.get("Protocol+SAP / SAP Link") or "").strip()
    if not protocol and not sap:
        print(f"[{num}] skip: no links")
        return []
    if not slug:
        print(f"[{num}] skip: no usable Paper Link for folder name")
        return []

    row_dir = DOCS_DIR / slug
    row_dir.mkdir(parents=True, exist_ok=True)
    new_files: list[Path] = []

    if protocol:
        dest = row_dir / "protocol.pdf"
        existed = dest.exists()
        ok, msg = download_pdf(protocol, dest)
        print(f"[{num}/{slug}] protocol: {msg}")
        if ok and not existed:
            new_files.append(dest)
        elif not ok:
            (row_dir / "protocol.error.txt").write_text(
                f"{protocol}\n{msg}\n", encoding="utf-8"
            )

    if sap:
        dest = row_dir / "sap.pdf"
        existed = dest.exists()
        ok, msg = download_pdf(sap, dest)
        print(f"[{num}/{slug}] sap: {msg}")
        if ok and not existed:
            new_files.append(dest)
        elif not ok:
            (row_dir / "sap.error.txt").write_text(
                f"{sap}\n{msg}\n", encoding="utf-8"
            )

    return new_files


def _hf_cli() -> str:
    # Resolve the `hf` CLI even when PATH doesn't include the active venv/conda env
    # (e.g. when this script is launched from a non-interactive shell).
    import shutil

    found = shutil.which("hf")
    if found:
        return found
    candidate = Path(sys.executable).parent / "hf"
    if candidate.exists():
        return str(candidate)
    raise RuntimeError(
        "`hf` CLI not found. Install with: pip install -U 'huggingface_hub[cli]'"
    )


def hf_upload(paths: list[Path]) -> None:
    if not paths:
        print("No files to upload.")
        return
    rels = [str(p.relative_to(ROOT)) for p in paths]
    print(f"Uploading {len(rels)} files to {HF_REPO} ...")
    hf = _hf_cli()
    # Use hf upload for small incremental batches; switch to upload-large-folder
    # if the new-row set is large.
    if len(rels) > 200:
        cmd = [
            hf, "upload-large-folder", HF_REPO, str(ROOT),
            "--repo-type=dataset", "--num-workers=4",
        ]
        subprocess.run(cmd, check=True)
        return
    for rel in rels:
        cmd = [hf, "upload", HF_REPO, rel, rel, "--repo-type=dataset"]
        subprocess.run(cmd, check=True)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--no-upload", action="store_true", help="Skip HF upload.")
    parser.add_argument(
        "--dry-run", action="store_true", help="Show diff only, no downloads or upload."
    )
    parser.add_argument(
        "--data-dir",
        default=None,
        help="Directory containing data/ and documents/. Defaults to the script directory.",
    )
    args = parser.parse_args()

    if args.data_dir:
        _set_data_dir(Path(args.data_dir))
    print(f"Data directory: {ROOT}")

    print(f"Fetching sheet ({SHEET_ID}, gid={SHEET_GID}) ...")
    csv_text = fetch_sheet_csv()
    new_rows_all = parse_rows(csv_text)
    print(f"Sheet has {len(new_rows_all)} data rows.")

    existing_ids = read_existing_ids()
    print(f"Local parquet has {len(existing_ids)} rows.")

    new_in_sheet = [r for r in new_rows_all if (r.get("#") or "").strip() not in existing_ids]
    rows_needing_pdfs = [r for r in new_rows_all if _row_missing_pdfs(r)]
    print(f"New rows in sheet (vs parquet #): {len(new_in_sheet)}")
    print(f"Rows missing PDFs on disk: {len(rows_needing_pdfs)}")
    for r in rows_needing_pdfs[:20]:
        print(f"  - #{r.get('#')}: {(r.get('Paper Title') or '')[:80]}")
    if len(rows_needing_pdfs) > 20:
        print(f"  ... and {len(rows_needing_pdfs) - 20} more")

    if args.dry_run:
        return

    write_parquet(csv_text)
    print(f"Wrote {PARQUET_PATH}")

    DOCS_DIR.mkdir(exist_ok=True)
    new_pdfs: list[Path] = []
    for row in rows_needing_pdfs:
        new_pdfs.extend(download_for_row(row))

    print(f"Downloaded {len(new_pdfs)} new PDFs.")

    if args.no_upload:
        return

    hf_upload([PARQUET_PATH, *new_pdfs])
    print("Done.")


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\nAborted.")
        sys.exit(130)
